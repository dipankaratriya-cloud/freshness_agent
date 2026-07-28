"""
Compare a provenance_refresh_extractor.py results JSON against the
ground-truth evidence dates in provenance_evidence_dates.xlsx.

Ground truth is keyed by provenance_url and may contain multiple rows per
URL (a landing page can cover several underlying datasets, each with its
own evidence-derived date). A pipeline date counts as a match if it equals
ANY of the ground-truth dates recorded for that URL.

Buckets (deduplicated by URL):
  match         - pipeline date equals one of the ground-truth dates
  mismatch      - pipeline found a date, ground truth has date(s), none match
  miss          - pipeline found no date, ground truth has date(s)
  false_pos     - pipeline found a date, ground truth has no date for this URL
  correct_null  - pipeline found no date, ground truth has no date for this URL
  no_ground_truth - URL not present in the golden set at all (not scored)

Usage:
  python3 validate.py --results provenance_top50_v4_results.json \
                       --ground-truth provenance_evidence_dates.xlsx
"""

import argparse
import json
from collections import defaultdict

import openpyxl


def load_ground_truth(xlsx_path: str) -> dict[str, list[str | None]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Evidence Dates"] if "Evidence Dates" in wb.sheetnames else wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    url_i = header.index("provenance_url")
    date_i = header.index("last_refresh_date")

    by_url: dict[str, list[str | None]] = defaultdict(list)
    for r in rows[1:]:
        if not r or not r[url_i]:
            continue
        url = str(r[url_i]).strip()
        val = r[date_i]
        date_str = str(val).strip()[:10] if val else None
        by_url[url].append(date_str)
    return by_url


def load_results(results_path: str) -> dict[str, str | None]:
    """Returns {url: last_refresh_date_or_None}, deduplicated by URL."""
    with open(results_path) as f:
        data = json.load(f)
    by_url: dict[str, str | None] = {}
    for entry in data.values():
        url = entry.get("url")
        if not url:
            continue
        by_url[url] = entry.get("last_refresh_date")
    return by_url


def validate(results_path: str, ground_truth_path: str, verbose: bool = False):
    gt = load_ground_truth(ground_truth_path)
    pred = load_results(results_path)

    buckets = defaultdict(list)
    for url, pred_date in pred.items():
        if url not in gt:
            buckets["no_ground_truth"].append((url, pred_date, None))
            continue
        gt_dates = [d for d in gt[url] if d]
        if pred_date:
            if gt_dates and pred_date in gt_dates:
                buckets["match"].append((url, pred_date, gt_dates))
            elif gt_dates:
                buckets["mismatch"].append((url, pred_date, gt_dates))
            else:
                buckets["false_pos"].append((url, pred_date, gt_dates))
        else:
            if gt_dates:
                buckets["miss"].append((url, pred_date, gt_dates))
            else:
                buckets["correct_null"].append((url, pred_date, gt_dates))

    scored = sum(len(buckets[k]) for k in
                 ("match", "mismatch", "miss", "false_pos", "correct_null"))

    print(f"Results file   : {results_path}")
    print(f"Ground truth   : {ground_truth_path}")
    print(f"Predicted URLs : {len(pred)}  (scored against golden set: {scored})")
    print("─" * 50)
    for key in ("match", "mismatch", "miss", "false_pos", "correct_null", "no_ground_truth"):
        print(f"  {key:15s}: {len(buckets[key])}")
    denom = scored - len(buckets["correct_null"]) - len(buckets["no_ground_truth"])
    comparable = len(buckets["match"]) + len(buckets["mismatch"])
    if comparable:
        print(f"\nAccuracy (match / (match+mismatch)) = "
              f"{len(buckets['match'])}/{comparable} = "
              f"{len(buckets['match'])*100//comparable}%")

    if verbose:
        for key in ("mismatch", "miss", "false_pos"):
            if not buckets[key]:
                continue
            print(f"\n── {key} ──")
            for url, pred_date, gt_dates in buckets[key]:
                print(f"  {url[:70]:70s}  pred={pred_date or '—':12s}  gt={gt_dates}")

    return buckets


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--results", required=True)
    ap.add_argument("--ground-truth", default="provenance_evidence_dates.xlsx")
    ap.add_argument("--verbose", "-v", action="store_true",
                     help="list URLs in mismatch/miss/false_pos buckets")
    args = ap.parse_args()
    validate(args.results, args.ground_truth, verbose=args.verbose)
